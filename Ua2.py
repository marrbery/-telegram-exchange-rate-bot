import requests
from bs4 import BeautifulSoup
import datetime
import telebot
import openpyxl
import os
import io

# Токен Telegram бота
BOT_TOKEN = ""

# Создание экземпляра бота
bot = telebot.TeleBot(BOT_TOKEN)

# Функція для отримання курсу долара до гривні з веб-сторінки
def get_exchange_rate():
    url = "https://www.google.com/finance/quote/USD-UAH"
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        exchange_rate_element = soup.find("div", class_="YMlKec fxKbKc")
        if exchange_rate_element:
            exchange_rate = exchange_rate_element.text.strip()
            return exchange_rate
    return None

# Обробник команди /get_exchange_rate
@bot.message_handler(commands=['get_exchange_rate'])
def send_exchange_rate(message):
    # Отримання поточної дати та часу
    now = datetime.datetime.now()
    current_time = now.strftime('%Y-%m-%d %H:%M:%S')

    # Отримання курсу валют
    exchange_rate = get_exchange_rate()

    if exchange_rate:
        # Створення xlsx-файлу
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Додавання заголовка
        worksheet['A1'] = 'Час'
        worksheet['B1'] = 'Курс'

        # Додавання даних
        worksheet.cell(row=2, column=1, value=current_time)
        worksheet.cell(row=2, column=2, value=exchange_rate)

        # Збереження файлу у буфері
        file_bytes = io.BytesIO()
        workbook.save(file_bytes)
        file_bytes.seek(0)

        # Відправка файлу користувачеві
        file_name = f'exchange_rate_{now.date()}.xlsx'
        bot.send_document(chat_id=message.chat.id, document=file_bytes, caption='Курс долара до гривні', visible_file_name=file_name)
    else:
        bot.reply_to(message, 'Не вдалося отримати курс валют')

# Запуск бота
bot.polling()
