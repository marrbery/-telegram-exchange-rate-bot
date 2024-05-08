import os
import io
import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
import telebot
import sqlite3

# Завантаження налаштувань з змінних оточення
BOT_TOKEN = os.environ.get('BOT_TOKEN', 'YOUR_BOT_TOKEN_HERE')
if not BOT_TOKEN:
    raise ValueError('BOT_TOKEN environment variable not set')

FINANCE_URL = 'https://www.google.com/finance/quote/USD-UAH'

# Створення екземпляру бота
bot = telebot.TeleBot(BOT_TOKEN)

# Підключення до бази даних
conn = sqlite3.connect('exchange_rates.db')
c = conn.cursor()
c.execute('''CREATE TABLE IF NOT EXISTS exchange_rates
             (timestamp TEXT, rate REAL)''')
conn.commit()

# Функція для зберігання курсу валют у базі даних
def save_exchange_rate(timestamp, rate):
    c.execute("INSERT INTO exchange_rates VALUES (?, ?)", (timestamp, rate))
    conn.commit()

# Функція для отримання даних з бази даних
def get_exchange_rates_from_db(date):
    c.execute("SELECT timestamp, rate FROM exchange_rates WHERE date(timestamp) = ?", (date,))
    return c.fetchall()

def get_exchange_rate():
    """
    Отримує курс долара США до гривні з веб-сайту Google Finance
    Зберігає дані у базі даних та повертає рядок з курсом або None у випадку помилки
    """
    try:
        response = requests.get(FINANCE_URL)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        rate_element = soup.select_one('div.BNeawe.iBp4i.AP7Wnd > div.YMlKec.fxKbKc')
        if rate_element:
            rate = rate_element.text.strip()
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            save_exchange_rate(timestamp, rate)
            return rate
    except requests.exceptions.RequestException as e:
        print(f'Помилка отримання курсу валют: {e}')
    return None

@bot.message_handler(commands=['get_exchange_rate'])
def send_exchange_rate(message):
    """
    Обробник команди /get_exchange_rate у Telegram
    Отримує курс долара США до гривні та відправляє його у файлі Excel
    """
    now = datetime.datetime.now()
    rate = get_exchange_rate() # Отримання поточного курсу валют
    if rate:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Час'
        worksheet['B1'] = 'Курс валют'
        row = 2
        # Додавання даних з бази даних до файлу Excel
        for timestamp, rate in get_exchange_rates_from_db(now.date()):
            worksheet.cell(row=row, column=1, value=timestamp)
            worksheet.cell(row=row, column=2, value=rate)
            row += 1
        file_bytes = io.BytesIO()
        workbook.save(file_bytes)
        file_bytes.seek(0)
        file_name = f'exchange_rate_{now.date()}.xlsx'
        bot.send_document(message.chat.id, file_bytes, caption='Курс долара США до гривні', visible_file_name=file_name)
    else:
        bot.reply_to(message, 'Не вдалося отримати курс валют')

if __name__ == '__main__':
    print('Запуск бота...')
    bot.polling()
